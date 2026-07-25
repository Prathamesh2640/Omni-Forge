"""Unit tests for the document_suite logic layer.

Conversions are exercised against real files and the output is parsed back,
so a malformed result fails rather than merely being written.
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
from pathlib import Path

import pytest
import yaml

from core.models import ProgressEvent
from modules.converters.document_suite.constants import (
    OUTPUT_SUBDIR,
    PROGRESS_COMPLETE,
)
from modules.converters.document_suite.logic import DocumentSuiteLogic
from modules.converters.document_suite.models import (
    Conversion,
    ConvertParams,
    ConvertResult,
)


@pytest.fixture()
def logic() -> DocumentSuiteLogic:
    """Fresh logic instance with no EventBus registrations."""
    return DocumentSuiteLogic()


@pytest.fixture()
def out_dir(tmp_path: Path) -> Path:
    """Directory conversions write into."""
    return tmp_path / "exports"


def write(path: Path, content: str) -> Path:
    """Write text to *path* and return it."""
    path.write_text(content, encoding="utf-8")
    return path


def run(
    logic: DocumentSuiteLogic, params: ConvertParams
) -> tuple[list[ProgressEvent], ConvertResult]:
    """Execute a conversion to completion.

    Returns:
        The emitted events and the recorded result.
    """

    async def drive() -> list[ProgressEvent]:
        return [event async for event in logic.execute(params)]

    events = asyncio.run(drive())
    assert logic._last_result is not None
    return events, logic._last_result


def _params(
    conversion: Conversion, inputs: list[Path], out: Path, **kw: object
) -> ConvertParams:
    """Build ConvertParams with the common fields filled in."""
    return ConvertParams(
        conversion=conversion, input_paths=inputs, output_dir=out, **kw
    )  # type: ignore[arg-type]


# ─── Markdown / HTML ──────────────────────────────────────────────────────────


class TestMarkdownToHtml:
    def test_renders_headings_and_emphasis(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "doc.md", "# Title\n\nSome **bold** text.\n")
        _events, result = run(
            logic, _params(Conversion.MARKDOWN_TO_HTML, [source], out_dir)
        )

        html = result.output_paths[0].read_text(encoding="utf-8")
        assert "<h1>Title</h1>" in html
        assert "<strong>bold</strong>" in html

    def test_output_is_self_contained(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        """No CDN or external stylesheet, so it renders offline (rule C-01)."""
        source = write(tmp_path / "doc.md", "# Title\n\n```python\nx = 1\n```\n")
        _events, result = run(
            logic, _params(Conversion.MARKDOWN_TO_HTML, [source], out_dir)
        )

        html = result.output_paths[0].read_text(encoding="utf-8")
        assert "<style>" in html
        assert "http://" not in html
        assert "https://" not in html

    def test_tables_are_rendered(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(
            tmp_path / "doc.md", "| A | B |\n|---|---|\n| 1 | 2 |\n"
        )
        _events, result = run(
            logic, _params(Conversion.MARKDOWN_TO_HTML, [source], out_dir)
        )
        assert "<table>" in result.output_paths[0].read_text(encoding="utf-8")

    def test_writes_into_the_module_subdirectory(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "doc.md", "# Hi\n")
        _events, result = run(
            logic, _params(Conversion.MARKDOWN_TO_HTML, [source], out_dir)
        )
        assert result.output_paths[0].parent.name == OUTPUT_SUBDIR


class TestHtmlToMarkdown:
    def test_converts_headings_and_links(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(
            tmp_path / "page.html",
            "<h1>Title</h1><p>See <a href='https://x.test'>here</a>.</p>",
        )
        _events, result = run(
            logic, _params(Conversion.HTML_TO_MARKDOWN, [source], out_dir)
        )

        markdown = result.output_paths[0].read_text(encoding="utf-8")
        assert "# Title" in markdown
        assert "[here](https://x.test)" in markdown

    def test_round_trips_back_to_html(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        """Markdown produced here must be readable by the forward conversion."""
        source = write(tmp_path / "page.html", "<h2>Section</h2><p>Body text.</p>")
        _events, first = run(
            logic, _params(Conversion.HTML_TO_MARKDOWN, [source], out_dir)
        )

        _events, second = run(
            logic, _params(Conversion.MARKDOWN_TO_HTML, [first.output_paths[0]], out_dir)
        )
        assert "<h2>Section</h2>" in second.output_paths[0].read_text(encoding="utf-8")


# ─── JSON / CSV ───────────────────────────────────────────────────────────────


class TestJsonToCsv:
    def test_writes_a_header_and_rows(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(
            tmp_path / "data.json",
            json.dumps([{"name": "Ada", "age": 36}, {"name": "Alan", "age": 41}]),
        )
        _events, result = run(logic, _params(Conversion.JSON_TO_CSV, [source], out_dir))

        rows = list(csv.DictReader(io.StringIO(
            result.output_paths[0].read_text(encoding="utf-8")
        )))
        assert [r["name"] for r in rows] == ["Ada", "Alan"]
        assert result.records == 2

    def test_nested_objects_become_dotted_columns(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(
            tmp_path / "data.json",
            json.dumps([{"user": {"city": "Pune", "zip": "411001"}}]),
        )
        _events, result = run(logic, _params(Conversion.JSON_TO_CSV, [source], out_dir))

        text = result.output_paths[0].read_text(encoding="utf-8")
        assert "user.city" in text
        assert "Pune" in text

    def test_flattening_can_be_disabled(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "data.json", json.dumps([{"user": {"city": "Pune"}}]))
        _events, result = run(
            logic,
            _params(Conversion.JSON_TO_CSV, [source], out_dir, flatten_nested=False),
        )

        text = result.output_paths[0].read_text(encoding="utf-8")
        assert "user.city" not in text
        assert "user" in text

    def test_a_single_object_becomes_one_row(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "data.json", json.dumps({"name": "Ada"}))
        _events, result = run(logic, _params(Conversion.JSON_TO_CSV, [source], out_dir))
        assert result.records == 1

    def test_rows_with_different_keys_share_a_header(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        """A ragged array must not lose columns that only later rows have."""
        source = write(
            tmp_path / "data.json", json.dumps([{"a": 1}, {"b": 2}])
        )
        _events, result = run(logic, _params(Conversion.JSON_TO_CSV, [source], out_dir))

        header = result.output_paths[0].read_text(encoding="utf-8").splitlines()[0]
        assert "a" in header
        assert "b" in header

    def test_lists_are_kept_as_json_text(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "data.json", json.dumps([{"tags": ["x", "y"]}]))
        _events, result = run(logic, _params(Conversion.JSON_TO_CSV, [source], out_dir))

        # Read it back through the CSV parser: the cell is quoted on disk.
        rows = list(csv.DictReader(io.StringIO(
            result.output_paths[0].read_text(encoding="utf-8")
        )))
        assert json.loads(rows[0]["tags"]) == ["x", "y"]

    def test_a_top_level_scalar_is_rejected(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "data.json", json.dumps(42))
        with pytest.raises(ValueError, match="object or an array of objects"):
            run(logic, _params(Conversion.JSON_TO_CSV, [source], out_dir))

    def test_a_non_object_item_names_its_position(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "data.json", json.dumps([{"a": 1}, "oops"]))
        with pytest.raises(ValueError, match="item 2 is a str"):
            run(logic, _params(Conversion.JSON_TO_CSV, [source], out_dir))


class TestCsvToJson:
    def test_reads_rows_into_records(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "data.csv", "name,age\nAda,36\nAlan,41\n")
        _events, result = run(logic, _params(Conversion.CSV_TO_JSON, [source], out_dir))

        records = json.loads(result.output_paths[0].read_text(encoding="utf-8"))
        assert records == [{"name": "Ada", "age": 36}, {"name": "Alan", "age": 41}]

    @pytest.mark.parametrize(
        ("raw", "expected"),
        [
            ("42", 42),
            ("-7", -7),
            ("3.14", 3.14),
            ("true", True),
            ("FALSE", False),
            ("yes", True),
            ("", None),
            ("null", None),
            ("hello", "hello"),
        ],
    )
    def test_infers_scalar_types(
        self,
        logic: DocumentSuiteLogic,
        tmp_path: Path,
        out_dir: Path,
        raw: str,
        expected: object,
    ) -> None:
        # A second column keeps the row non-blank, so an empty first value is
        # still a real record rather than a line csv.DictReader skips.
        source = write(tmp_path / "data.csv", f"value,marker\n{raw},x\n")
        _events, result = run(logic, _params(Conversion.CSV_TO_JSON, [source], out_dir))

        records = json.loads(result.output_paths[0].read_text(encoding="utf-8"))
        assert records[0]["value"] == expected

    def test_inference_can_be_disabled(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        """Leading zeros must survive when the user asks for raw strings."""
        source = write(tmp_path / "data.csv", "code\n007\n")
        _events, result = run(
            logic, _params(Conversion.CSV_TO_JSON, [source], out_dir, infer_types=False)
        )

        records = json.loads(result.output_paths[0].read_text(encoding="utf-8"))
        assert records[0]["code"] == "007"

    def test_tsv_uses_tab_delimiters(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "data.tsv", "name\tage\nAda\t36\n")
        _events, result = run(logic, _params(Conversion.CSV_TO_JSON, [source], out_dir))

        records = json.loads(result.output_paths[0].read_text(encoding="utf-8"))
        assert records == [{"name": "Ada", "age": 36}]

    def test_a_byte_order_mark_is_stripped(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        """Excel writes UTF-8 CSVs with a BOM, which would corrupt the first key."""
        source = tmp_path / "data.csv"
        source.write_bytes(b"\xef\xbb\xbfname,age\nAda,36\n")
        _events, result = run(logic, _params(Conversion.CSV_TO_JSON, [source], out_dir))

        records = json.loads(result.output_paths[0].read_text(encoding="utf-8"))
        assert "name" in records[0]

    def test_an_empty_file_yields_no_records(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "data.csv", "name,age\n")
        _events, result = run(logic, _params(Conversion.CSV_TO_JSON, [source], out_dir))
        assert result.records == 0


# ─── Excel ────────────────────────────────────────────────────────────────────


class TestJsonToExcel:
    def test_writes_a_readable_workbook(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        from openpyxl import load_workbook

        source = write(
            tmp_path / "data.json",
            json.dumps([{"name": "Ada", "age": 36}, {"name": "Alan", "age": 41}]),
        )
        _events, result = run(
            logic, _params(Conversion.JSON_TO_EXCEL, [source], out_dir)
        )

        sheet = load_workbook(result.output_paths[0]).active
        assert sheet is not None
        assert [c.value for c in sheet[1]] == ["name", "age"]
        assert [c.value for c in sheet[2]] == ["Ada", 36]

    def test_the_header_row_is_bold(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        from openpyxl import load_workbook

        source = write(tmp_path / "data.json", json.dumps([{"name": "Ada"}]))
        _events, result = run(
            logic, _params(Conversion.JSON_TO_EXCEL, [source], out_dir)
        )

        sheet = load_workbook(result.output_paths[0]).active
        assert sheet is not None
        assert sheet["A1"].font.bold is True

    def test_nested_values_are_stored_as_json_text(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        from openpyxl import load_workbook

        source = write(
            tmp_path / "data.json", json.dumps([{"tags": ["a", "b"]}])
        )
        _events, result = run(
            logic, _params(Conversion.JSON_TO_EXCEL, [source], out_dir)
        )

        sheet = load_workbook(result.output_paths[0]).active
        assert sheet is not None
        assert sheet["A2"].value == '["a", "b"]'

    def test_columns_are_widened_within_bounds(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        from openpyxl import load_workbook

        from modules.converters.document_suite.constants import (
            EXCEL_MAX_COLUMN_WIDTH,
            EXCEL_MIN_COLUMN_WIDTH,
        )

        source = write(
            tmp_path / "data.json", json.dumps([{"a": "x", "b": "y" * 500}])
        )
        _events, result = run(
            logic, _params(Conversion.JSON_TO_EXCEL, [source], out_dir)
        )

        dimensions = load_workbook(result.output_paths[0]).active.column_dimensions
        assert dimensions["A"].width >= EXCEL_MIN_COLUMN_WIDTH
        assert dimensions["B"].width == EXCEL_MAX_COLUMN_WIDTH


# ─── JSON / YAML ──────────────────────────────────────────────────────────────


class TestYamlConversions:
    def test_json_becomes_yaml(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(
            tmp_path / "config.json", json.dumps({"name": "omniforge", "port": 8765})
        )
        _events, result = run(
            logic, _params(Conversion.JSON_TO_YAML, [source], out_dir)
        )

        loaded = yaml.safe_load(result.output_paths[0].read_text(encoding="utf-8"))
        assert loaded == {"name": "omniforge", "port": 8765}

    def test_yaml_becomes_json(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "config.yaml", "name: omniforge\nport: 8765\n")
        _events, result = run(
            logic, _params(Conversion.YAML_TO_JSON, [source], out_dir)
        )

        loaded = json.loads(result.output_paths[0].read_text(encoding="utf-8"))
        assert loaded == {"name": "omniforge", "port": 8765}

    def test_key_order_is_preserved(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        """Alphabetising a config file would make diffs unreadable."""
        source = write(
            tmp_path / "config.json", json.dumps({"zebra": 1, "apple": 2, "mango": 3})
        )
        _events, result = run(
            logic, _params(Conversion.JSON_TO_YAML, [source], out_dir)
        )

        text = result.output_paths[0].read_text(encoding="utf-8")
        assert text.index("zebra") < text.index("apple") < text.index("mango")

    def test_nested_structures_survive_a_round_trip(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        original = {"a": {"b": [1, 2, {"c": True}]}, "d": None}
        source = write(tmp_path / "config.json", json.dumps(original))

        _events, to_yaml = run(
            logic, _params(Conversion.JSON_TO_YAML, [source], out_dir)
        )
        renamed = to_yaml.output_paths[0].with_name("roundtrip.yaml")
        renamed.write_text(
            to_yaml.output_paths[0].read_text(encoding="utf-8"), encoding="utf-8"
        )
        _events, back = run(
            logic, _params(Conversion.YAML_TO_JSON, [renamed], out_dir)
        )

        assert json.loads(back.output_paths[0].read_text(encoding="utf-8")) == original

    def test_unicode_is_preserved(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(
            tmp_path / "config.json",
            json.dumps({"city": "पुणे"}, ensure_ascii=False),

        )
        _events, result = run(
            logic, _params(Conversion.JSON_TO_YAML, [source], out_dir)
        )
        assert "पुणे" in result.output_paths[0].read_text(encoding="utf-8")

    def test_yaml_loading_is_safe(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        """safe_load must refuse to instantiate arbitrary Python objects."""
        source = write(
            tmp_path / "evil.yaml", "!!python/object/apply:os.system ['echo pwned']\n"
        )
        with pytest.raises(yaml.YAMLError):
            run(logic, _params(Conversion.YAML_TO_JSON, [source], out_dir))


# ─── Shared behaviour ─────────────────────────────────────────────────────────


class TestExecutionContract:
    def test_every_conversion_has_an_implementation(
        self, logic: DocumentSuiteLogic
    ) -> None:
        assert set(logic._converters()) == set(Conversion)

    def test_a_missing_input_is_reported_clearly(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        with pytest.raises(FileNotFoundError, match="File not found"):
            run(
                logic,
                _params(Conversion.JSON_TO_YAML, [tmp_path / "absent.json"], out_dir),
            )

    def test_a_mismatched_extension_is_rejected(
        self, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "data.csv", "a,b\n1,2\n")
        with pytest.raises(ValueError, match="requires"):
            _params(Conversion.JSON_TO_YAML, [source], out_dir)

    def test_several_files_are_converted_independently(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        sources = [
            write(tmp_path / f"doc{i}.md", f"# Doc {i}\n") for i in range(3)
        ]
        _events, result = run(
            logic, _params(Conversion.MARKDOWN_TO_HTML, sources, out_dir)
        )

        assert len(result.output_paths) == 3
        assert all(p.is_file() for p in result.output_paths)

    def test_progress_ends_at_one_hundred(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "doc.md", "# Hi\n")
        events, _result = run(
            logic, _params(Conversion.MARKDOWN_TO_HTML, [source], out_dir)
        )
        assert events[-1].percent == PROGRESS_COMPLETE

    def test_the_final_event_carries_the_output_path(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "doc.md", "# Hi\n")
        events, _result = run(
            logic, _params(Conversion.MARKDOWN_TO_HTML, [source], out_dir)
        )
        assert events[-1].output_path is not None

    def test_sizes_are_recorded(
        self, logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
    ) -> None:
        source = write(tmp_path / "doc.md", "# Hi\n")
        _events, result = run(
            logic, _params(Conversion.MARKDOWN_TO_HTML, [source], out_dir)
        )
        assert result.input_bytes > 0
        assert result.output_bytes > 0


# ─── EventBus wiring ──────────────────────────────────────────────────────────


@pytest.mark.asyncio()
async def test_register_and_unregister_round_trip(logic: DocumentSuiteLogic) -> None:
    from core.event_bus import event_bus
    from modules.converters.document_suite.constants import EVENT_EXECUTE

    await logic.register()
    assert logic._on_execute in event_bus._subscribers[EVENT_EXECUTE]

    await logic.unregister()
    assert logic._on_execute not in event_bus._subscribers[EVENT_EXECUTE]


@pytest.mark.asyncio()
async def test_result_is_published_on_completion(
    logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
) -> None:
    from core.event_bus import event_bus
    from modules.converters.document_suite.constants import EVENT_DONE

    received: list[ConvertResult] = []

    async def capture(payload: object) -> None:
        assert isinstance(payload, ConvertResult)
        received.append(payload)

    event_bus.subscribe(EVENT_DONE, capture)
    try:
        source = write(tmp_path / "doc.md", "# Hi\n")
        await logic._on_execute(
            _params(Conversion.MARKDOWN_TO_HTML, [source], out_dir)
        )
    finally:
        event_bus.unsubscribe(EVENT_DONE, capture)

    assert len(received) == 1


@pytest.mark.asyncio()
async def test_a_failure_publishes_an_error_not_a_result(
    logic: DocumentSuiteLogic, tmp_path: Path, out_dir: Path
) -> None:
    from core.event_bus import event_bus
    from modules.converters.document_suite.constants import EVENT_DONE, EVENT_ERROR

    done: list[object] = []
    errors: list[object] = []

    async def on_done(payload: object) -> None:
        done.append(payload)

    async def on_error(payload: object) -> None:
        errors.append(payload)

    event_bus.subscribe(EVENT_DONE, on_done)
    event_bus.subscribe(EVENT_ERROR, on_error)
    try:
        await logic._on_execute(
            _params(Conversion.JSON_TO_YAML, [tmp_path / "absent.json"], out_dir)
        )
    finally:
        event_bus.unsubscribe(EVENT_DONE, on_done)
        event_bus.unsubscribe(EVENT_ERROR, on_error)

    assert done == []
    assert len(errors) == 1


@pytest.mark.asyncio()
async def test_a_non_params_payload_is_ignored(logic: DocumentSuiteLogic) -> None:
    await logic._on_execute({"conversion": "json_to_yaml"})
    assert logic._last_result is None


# ─── Helpers ──────────────────────────────────────────────────────────────────


class TestHelpers:
    @pytest.mark.parametrize(
        ("value", "expected"),
        [(None, None), ("plain", "plain"), (42, 42), (True, True), (1.5, 1.5)],
    )
    def test_scalars_pass_into_excel_unchanged(
        self, logic: DocumentSuiteLogic, value: object, expected: object
    ) -> None:
        assert logic._excel_value(value) == expected

    @pytest.mark.parametrize(
        ("value", "expected"),
        [({"a": 1}, '{"a": 1}'), ([1, 2], "[1, 2]")],
    )
    def test_containers_become_json_text_for_excel(
        self, logic: DocumentSuiteLogic, value: object, expected: str
    ) -> None:
        """openpyxl cannot store a dict or list in a cell."""
        assert logic._excel_value(value) == expected

    def test_a_missing_csv_value_infers_as_null(
        self, logic: DocumentSuiteLogic
    ) -> None:
        """DictReader yields None for columns a short row does not reach."""
        assert logic._infer(None) is None

    @pytest.mark.parametrize(
        ("data", "expected"),
        [([1, 2, 3], 3), ({"a": 1, "b": 2}, 2), ("scalar", 1), (42, 1), (None, 1)],
    )
    def test_counts_top_level_items(
        self, logic: DocumentSuiteLogic, data: object, expected: int
    ) -> None:
        assert logic._count_items(data) == expected

    def test_flatten_handles_deep_nesting(self, logic: DocumentSuiteLogic) -> None:
        flat = logic.flatten({"a": {"b": {"c": 1}}})
        assert flat == {"a.b.c": 1}

    def test_flatten_leaves_scalars_alone(self, logic: DocumentSuiteLogic) -> None:
        assert logic.flatten({"x": 1, "y": None}) == {"x": 1, "y": None}
