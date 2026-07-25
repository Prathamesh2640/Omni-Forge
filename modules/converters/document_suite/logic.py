"""Document Suite — business logic layer.

Every conversion runs through pure-Python libraries (mistune, PyYAML,
openpyxl, markdownify), so nothing here needs an external binary or a network
connection (rule C-01).

Zero NiceGUI imports permitted (rule A-01).
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
from collections.abc import AsyncIterator, Callable
from pathlib import Path
from typing import Any

from core.event_bus import event_bus
from core.logger import get_logger
from core.models import ProgressEvent
from core.sandbox import SandboxTask, run_in_thread
from modules.converters.document_suite.constants import (
    CODE_HIGHLIGHT_STYLE,
    CSV_FALSE_VALUES,
    CSV_NULL_VALUES,
    CSV_TRUE_VALUES,
    EVENT_CANCEL,
    EVENT_CANCELLED,
    EVENT_DONE,
    EVENT_ERROR,
    EVENT_EXECUTE,
    EVENT_PROGRESS,
    EXCEL_MAX_COLUMN_WIDTH,
    EXCEL_MIN_COLUMN_WIDTH,
    EXCEL_SHEET_NAME,
    EXCEL_WIDTH_PADDING,
    FLATTEN_SEPARATOR,
    HTML_BASE_CSS,
    HTML_DOCUMENT_TEMPLATE,
    JSON_INDENT,
    OUTPUT_SUBDIR,
    PROGRESS_COMPLETE,
    PROGRESS_START,
    YAML_INDENT,
    YAML_LINE_WIDTH,
)
from modules.converters.document_suite.models import (
    OUTPUT_EXTENSION,
    Conversion,
    ConvertParams,
    ConvertResult,
)
from shared.constants import DEFAULT_EXECUTION_TIMEOUT_SECONDS
from shared.validators import validate_write_target

logger = get_logger(__name__)


class DocumentSuiteLogic:
    """Implements every document_suite conversion."""

    def __init__(self) -> None:
        self._execution = SandboxTask()
        self._last_result: ConvertResult | None = None

    async def register(self) -> None:
        """Subscribe the EventBus execute handler.  Call from ``on_load()``."""
        event_bus.subscribe(EVENT_EXECUTE, self._on_execute)
        event_bus.subscribe(EVENT_CANCEL, self._on_cancel)
        logger.debug("document_suite.logic.registered")

    async def unregister(self) -> None:
        """Unsubscribe the EventBus handler.  Call from ``on_unload()``."""
        event_bus.unsubscribe(EVENT_EXECUTE, self._on_execute)
        event_bus.unsubscribe(EVENT_CANCEL, self._on_cancel)
        logger.debug("document_suite.logic.unregistered")

    # ─── EventBus handler ─────────────────────────────────────────────────────

    async def _on_cancel(self, _payload: Any) -> None:
        """Stop the in-flight operation at the user's request.

        The cancelled run reports nothing itself — this handler owns telling
        the UI, so the execute handler can stay quiet about a deliberate
        user action (RFC 0003).
        """
        if self._execution.request_cancel():
            logger.info("document_suite.cancel_requested")
            await event_bus.publish(EVENT_CANCELLED, None)

    async def _on_execute(self, payload: Any) -> None:
        """Run a conversion requested by the UI.

        Args:
            payload: A ConvertParams instance.
        """
        if not isinstance(payload, ConvertParams):
            logger.error("document_suite.bad_payload — type=%s", type(payload).__name__)
            return

        self._last_result = None
        try:
            # Rule B-02 — bounded and cancellable. Iterating the generator here
            # directly meant the sandbox's timeout applied to nothing (RFC 0003).
            await self._execution.consume(
                self.execute(payload),
                lambda event: event_bus.publish(EVENT_PROGRESS, event),
            )
            if self._last_result is not None:
                await event_bus.publish(EVENT_DONE, self._last_result)
        except TimeoutError:
            logger.warning("document_suite.timeout — after %ds", DEFAULT_EXECUTION_TIMEOUT_SECONDS)
            await event_bus.publish(
                EVENT_ERROR,
                f"The operation exceeded {DEFAULT_EXECUTION_TIMEOUT_SECONDS}s and was stopped.",
            )
        except asyncio.CancelledError:
            # This handler task is itself the cancellation target and the cancel
            # handler has already told the UI, so a deliberate user action is
            # kept out of the error log.
            logger.info("document_suite.cancelled")
        except Exception as exc:
            logger.error("document_suite.execute_failed", exc_info=exc)
            await event_bus.publish(EVENT_ERROR, str(exc))

    # ─── Dispatch ─────────────────────────────────────────────────────────────

    async def execute(self, params: ConvertParams) -> AsyncIterator[ProgressEvent]:
        """Convert every input file, reporting progress throughout.

        Args:
            params: Validated conversion parameters.

        Yields:
            ProgressEvent at each checkpoint.
        """
        yield ProgressEvent(percent=PROGRESS_START, message="Validating input…")

        missing = [p for p in params.input_paths if not p.is_file()]
        if missing:
            raise FileNotFoundError(f"File not found: {missing[0]}")

        # Rule B-07 — confine writes to exports/, temp/, or the directory the
        # user chose for this run. Resolving here also stops a crafted filename
        # from escaping that directory via traversal.
        output_dir = validate_write_target(
            params.output_dir / OUTPUT_SUBDIR, extra_roots=(params.output_dir,)
        )
        output_dir.mkdir(parents=True, exist_ok=True)

        converter = self._converters()[params.conversion]
        outputs: list[Path] = []
        records = 0
        total = len(params.input_paths)

        for index, source in enumerate(params.input_paths, start=1):
            target = output_dir / (source.stem + OUTPUT_EXTENSION[params.conversion])
            produced, count = await run_in_thread(converter, source, target, params)
            outputs.append(produced)
            records += count
            yield ProgressEvent(
                percent=int(index / total * PROGRESS_COMPLETE),
                message=f"Converted {index}/{total}: {source.name}",
            )

        self._last_result = ConvertResult(
            conversion=params.conversion,
            output_paths=outputs,
            records=records,
            input_bytes=sum(p.stat().st_size for p in params.input_paths),
            output_bytes=sum(p.stat().st_size for p in outputs if p.is_file()),
            detail=self._summarise(params.conversion, outputs, records),
        )
        logger.info(
            "document_suite.done — conversion=%s files=%d records=%d",
            params.conversion.value,
            len(outputs),
            records,
        )
        yield ProgressEvent(
            percent=PROGRESS_COMPLETE,
            message=self._last_result.detail,
            output_path=outputs[0] if outputs else None,
        )

    def _converters(
        self,
    ) -> dict[Conversion, Callable[[Path, Path, ConvertParams], tuple[Path, int]]]:
        """Map each conversion to its implementation."""
        return {
            Conversion.MARKDOWN_TO_HTML: self._markdown_to_html,
            Conversion.HTML_TO_MARKDOWN: self._html_to_markdown,
            Conversion.JSON_TO_CSV: self._json_to_csv,
            Conversion.CSV_TO_JSON: self._csv_to_json,
            Conversion.JSON_TO_EXCEL: self._json_to_excel,
            Conversion.JSON_TO_YAML: self._json_to_yaml,
            Conversion.YAML_TO_JSON: self._yaml_to_json,
        }

    # ─── Markdown / HTML ──────────────────────────────────────────────────────

    def _markdown_to_html(
        self, source: Path, target: Path, _params: ConvertParams
    ) -> tuple[Path, int]:
        """Render Markdown to a self-contained HTML document."""
        import mistune
        from pygments.formatters import HtmlFormatter

        renderer = mistune.create_markdown(
            plugins=["table", "strikethrough", "task_lists", "url"],
            escape=False,
        )
        body = str(renderer(source.read_text(encoding="utf-8", errors="replace")))

        # types-Pygments leaves get_style_defs unannotated, so the call is
        # untyped even though the stub package is installed.
        formatter = HtmlFormatter(style=CODE_HIGHLIGHT_STYLE)
        code_css = str(formatter.get_style_defs(".highlight"))  # type: ignore[no-untyped-call]

        target.write_text(
            HTML_DOCUMENT_TEMPLATE.format(
                title=source.stem,
                base_css=HTML_BASE_CSS,
                code_css=code_css,
                body=body,
            ),
            encoding="utf-8",
        )
        return target, body.count("\n") + 1

    def _html_to_markdown(
        self, source: Path, target: Path, _params: ConvertParams
    ) -> tuple[Path, int]:
        """Convert an HTML document to Markdown."""
        from markdownify import markdownify

        markdown = str(
            markdownify(
                source.read_text(encoding="utf-8", errors="replace"),
                heading_style="ATX",
            )
        ).strip()
        target.write_text(markdown + "\n", encoding="utf-8")
        return target, markdown.count("\n") + 1

    # ─── JSON / CSV / Excel ───────────────────────────────────────────────────

    def _json_to_csv(
        self, source: Path, target: Path, params: ConvertParams
    ) -> tuple[Path, int]:
        """Write JSON records as CSV."""
        rows = self._tabular_rows(source, params)
        columns = self._columns(rows)

        with target.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return target, len(rows)

    def _json_to_excel(
        self, source: Path, target: Path, params: ConvertParams
    ) -> tuple[Path, int]:
        """Write JSON records into an .xlsx worksheet."""
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        rows = self._tabular_rows(source, params)
        columns = self._columns(rows)

        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = EXCEL_SHEET_NAME

        sheet.append(columns)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            sheet.append([self._excel_value(row.get(column)) for column in columns])

        self._autosize(sheet, columns, rows, get_column_letter)
        workbook.save(target)
        return target, len(rows)

    def _autosize(
        self,
        sheet: Any,
        columns: list[str],
        rows: list[dict[str, Any]],
        column_letter: Callable[[int], str],
    ) -> None:
        """Widen each column to fit its widest value, within bounds.

        Args:
            sheet: The worksheet being written.
            columns: Column headers in order.
            rows: The data rows.
            column_letter: openpyxl's index-to-letter helper.
        """
        for index, column in enumerate(columns, start=1):
            widest = max(
                [len(column)] + [len(str(row.get(column, ""))) for row in rows]
            )
            sheet.column_dimensions[column_letter(index)].width = min(
                max(widest + EXCEL_WIDTH_PADDING, EXCEL_MIN_COLUMN_WIDTH),
                EXCEL_MAX_COLUMN_WIDTH,
            )

    def _excel_value(self, value: Any) -> Any:
        """Convert a value into something openpyxl can store in a cell.

        Args:
            value: The value from the source record.

        Returns:
            The value itself for scalars, or its JSON text for containers.
        """
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        return json.dumps(value, ensure_ascii=False)

    def _csv_to_json(
        self, source: Path, target: Path, params: ConvertParams
    ) -> tuple[Path, int]:
        """Read CSV (or TSV) into JSON records, optionally inferring types."""
        text = source.read_text(encoding="utf-8-sig", errors="replace")
        delimiter = "\t" if source.suffix.lower() == ".tsv" else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

        records = [
            {
                key: (self._infer(value) if params.infer_types else value)
                for key, value in row.items()
                if key is not None
            }
            for row in reader
        ]

        target.write_text(
            json.dumps(records, indent=JSON_INDENT, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        return target, len(records)

    def _infer(self, value: str | None) -> Any:
        """Convert a CSV string into a number, boolean or null where obvious.

        Args:
            value: The raw cell text.

        Returns:
            The inferred value, or the original string when nothing matches.
        """
        if value is None:
            return None
        text = value.strip()
        lowered = text.lower()
        if lowered in CSV_NULL_VALUES:
            return None
        if lowered in CSV_TRUE_VALUES:
            return True
        if lowered in CSV_FALSE_VALUES:
            return False
        try:
            return int(text)
        except ValueError:
            pass
        try:
            return float(text)
        except ValueError:
            return value

    # ─── JSON / YAML ──────────────────────────────────────────────────────────

    def _json_to_yaml(
        self, source: Path, target: Path, _params: ConvertParams
    ) -> tuple[Path, int]:
        """Serialise a JSON document as YAML."""
        import yaml

        data = json.loads(source.read_text(encoding="utf-8"))
        target.write_text(
            yaml.safe_dump(
                data,
                indent=YAML_INDENT,
                width=YAML_LINE_WIDTH,
                sort_keys=False,
                allow_unicode=True,
            ),
            encoding="utf-8",
        )
        return target, self._count_items(data)

    def _yaml_to_json(
        self, source: Path, target: Path, _params: ConvertParams
    ) -> tuple[Path, int]:
        """Serialise a YAML document as JSON.

        Uses ``safe_load``, so a document cannot instantiate arbitrary Python
        objects on load.
        """
        import yaml

        data = yaml.safe_load(source.read_text(encoding="utf-8"))
        target.write_text(
            json.dumps(data, indent=JSON_INDENT, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        return target, self._count_items(data)

    # ─── Helpers ──────────────────────────────────────────────────────────────

    def _tabular_rows(self, source: Path, params: ConvertParams) -> list[dict[str, Any]]:
        """Load a JSON file as a list of flat records.

        Accepts either a top-level array of objects or a single object, which
        becomes a one-row table.

        Args:
            source: The JSON file.
            params: Conversion parameters, for the flattening choice.

        Returns:
            One dict per output row.

        Raises:
            ValueError: When the document is not an object or array of objects.
        """
        data = json.loads(source.read_text(encoding="utf-8"))

        if isinstance(data, dict):
            records: list[Any] = [data]
        elif isinstance(data, list):
            records = data
        else:
            raise ValueError(
                f"{source.name} must hold a JSON object or an array of objects."
            )

        rows: list[dict[str, Any]] = []
        for index, record in enumerate(records):
            if not isinstance(record, dict):
                raise ValueError(
                    f"{source.name}: item {index + 1} is a "
                    f"{type(record).__name__}, but every item must be an object."
                )
            rows.append(
                self.flatten(record) if params.flatten_nested else self._stringify(record)
            )
        return rows

    def flatten(
        self, record: dict[str, Any], prefix: str = ""
    ) -> dict[str, Any]:
        """Flatten nested objects into dotted column names.

        ``{"user": {"city": "Pune"}}`` becomes ``{"user.city": "Pune"}``.
        Lists are left as JSON text, since their length varies per row.

        Args:
            record: The record to flatten.
            prefix: Key prefix accumulated during recursion.

        Returns:
            A single-level mapping.
        """
        flat: dict[str, Any] = {}
        for key, value in record.items():
            column = f"{prefix}{key}"
            if isinstance(value, dict):
                flat.update(self.flatten(value, f"{column}{FLATTEN_SEPARATOR}"))
            elif isinstance(value, list):
                flat[column] = json.dumps(value, ensure_ascii=False)
            else:
                flat[column] = value
        return flat

    def _stringify(self, record: dict[str, Any]) -> dict[str, Any]:
        """Serialise nested values as JSON text, leaving the record flat.

        Args:
            record: The record to convert.

        Returns:
            A mapping whose values are all scalars or JSON strings.
        """
        return {
            key: (
                json.dumps(value, ensure_ascii=False)
                if isinstance(value, (dict, list))
                else value
            )
            for key, value in record.items()
        }

    def _columns(self, rows: list[dict[str, Any]]) -> list[str]:
        """Collect every column across all rows, in first-seen order.

        Args:
            rows: The records being written.

        Returns:
            Ordered column names.
        """
        columns: dict[str, None] = {}
        for row in rows:
            for key in row:
                columns.setdefault(key, None)
        return list(columns)

    def _count_items(self, data: Any) -> int:
        """Count the top-level entries of a parsed document.

        Args:
            data: The parsed document.

        Returns:
            Number of items, or 1 for a scalar.
        """
        if isinstance(data, (list, dict)):
            return len(data)
        return 1

    def _summarise(
        self, conversion: Conversion, outputs: list[Path], records: int
    ) -> str:
        """Build the completion message shown in the UI.

        Args:
            conversion: The conversion that ran.
            outputs: Files produced.
            records: Rows or items converted.

        Returns:
            A short human-readable summary.
        """
        noun = "file" if len(outputs) == 1 else "files"
        label = conversion.value.replace("_to_", " → ").replace("_", " ").title()
        return f"{label} complete — {len(outputs)} {noun}, {records:,} records"
